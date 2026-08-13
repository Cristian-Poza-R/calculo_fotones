import io
import os
import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from scipy.stats import linregress
import openpyxl
from openpyxl.chart import ScatterChart, Reference, Series

st.set_page_config(page_title="Calculadora de Flujo Fotónico", layout="wide")

# ESTRUCTURA DE LA BARRA LATERAL
st.sidebar.markdown(
    "<div style='text-align:center; font-size:14px; font-weight:bold; margin-bottom:10px;'>"
    "Laboratorio de Procesos Avanzados<br>"
    "de Oxidación y Química verde"
    "</div>", 
    unsafe_allow_html=True
)

# LOGO CENTRADO
logo_path = os.path.join(os.path.dirname(__file__), "logo.png")
col1, col_img, col3 = st.sidebar.columns([1, 4, 1])
with col_img:
    if os.path.exists(logo_path):
        st.image(logo_path, width=170)
    elif os.path.exists("logo.png"):
        st.image("logo.png", width=170)

st.sidebar.markdown(
    "<div style='text-align:center; font-size:14px; margin-top:5px;'>"
    "<b>Facultad de Ciencias Químicas</b><br>"
    "Universidad de Concepción - Chile."
    "</div>", 
    unsafe_allow_html=True
)

st.title("Calculadora de Flujo Fotónico")

# BARRA LATERAL: CONSTANTES Y VOLÚMENES EXPLICADOS
st.sidebar.markdown("---")
st.sidebar.header("Constantes y Volúmenes del Experimento")
st.sidebar.markdown(
    "<div style='font-size: 12px; color: #555; margin-bottom: 6px;'>"
    "<b>Algunas definiciones:</b><br>"
    "• <b>V₁:</b> Volumen del reactor (solución inicial de Fe³⁺).<br>"
    "• <b>V₂:</b> Volumen de alícuota tomada del reactor para el ensayo.<br>"
    "• <b>V₃:</b> Volumen final de enrase (formación del complejo Fe²⁺ con fenantrolina)."
    "</div>"
    "<div style='font-size: 12px; color: #555; margin-bottom: 10px;'>"
    "• <b>Φ(λ):</b> Rendimiento cuántico; puede variar según la lámpara utilizada."
    "</div>",
    unsafe_allow_html=True
)

v1 = st.sidebar.number_input("V1 (L)", value=0.0000419, format="%.8f")
v2 = st.sidebar.number_input("V2 (L)", value=0.0000419, format="%.8f")
v3 = st.sidebar.number_input("V3 (L)", value=0.0050000, format="%.8f")
epsilon = st.sidebar.number_input("Epsilon a 510 nm (M⁻¹cm⁻¹)", value=11100)
l = st.sidebar.number_input("Paso óptico l (cm)", value=1.0)
phi_lambda = st.sidebar.number_input("Φ(λ)", value=1.15)

v_total = v1 + v2 + v3

# DATOS INICIALES
tiempos_estandar = [0, 1, 3, 5, 10, 20, 30]
if "df" not in st.session_state:
    st.session_state.df = pd.DataFrame({
        "Tiempo (s)": tiempos_estandar,
        "Abs510nm-A": [0.0] * 7,
        "Abs510nm-B": [0.0] * 7,
        "Abs510nm-C": [0.0] * 7,
    })

st.subheader("1. Inserción de Datos Experimentales")
df_editado = st.data_editor(
    st.session_state.df, num_rows="fixed", use_container_width=True
)

# BARRA LATERAL: SELECCIÓN RÉPLICAS
st.sidebar.markdown("---")
st.sidebar.header("Filtro de Réplicas por Tiempo")
st.sidebar.write("Desmarca individualmente cualquier réplica defectuosa:")

activas = {}
for idx, fila in df_editado.iterrows():
    t = fila["Tiempo (s)"]
    st.sidebar.markdown(f"**Tiempo: {t} s**")
    c1, c2, c3 = st.sidebar.columns(3)
    activas[(t, "A")] = c1.checkbox("A", value=True, key=f"chk_A_{t}")
    activas[(t, "B")] = c2.checkbox("B", value=True, key=f"chk_B_{t}")
    activas[(t, "C")] = c3.checkbox("C", value=True, key=f"chk_C_{t}")

# CÁLCULOS DINÁMICOS FILTRANDO SÓLO FILAS CON RÉPLICAS ACTIVAS
def procesar_datos():
    resultados = []
    for _, fila in df_editado.iterrows():
        t = fila["Tiempo (s)"]
        vals = []

        if activas.get((t, "A"), True):
            vals.append(fila["Abs510nm-A"])
        if activas.get((t, "B"), True):
            vals.append(fila["Abs510nm-B"])
        if activas.get((t, "C"), True):
            vals.append(fila["Abs510nm-C"])

        if not vals:
            continue

        media = np.mean(vals)
        std_dev = np.std(vals, ddof=1) if len(vals) > 1 else 0.0

        moles_fe = (media * v_total) / (epsilon * l)
        moles_fe_std = (std_dev * v_total) / (epsilon * l)

        denominador = phi_lambda * epsilon * v2 * l * t
        q_np = (media * v1 * v3) / denominador if denominador > 0 and t > 0 else 0.0
        q_np_std = (std_dev * v1 * v3) / denominador if denominador > 0 and t > 0 else 0.0

        resultados.append(
            {
                "Tiempo (s)": t,
                "Media Abs": media,
                "Std Abs": std_dev,
                "moles Fe2+": moles_fe,
                "Std moles Fe2+": moles_fe_std,
                "q_n,p (einstein s⁻¹)": q_np,
                "Std q_n,p": q_np_std,
            }
        )
    return pd.DataFrame(resultados)

df_final = procesar_datos()

# 2. RESULTADOS CALCULADOS
st.markdown("---")
st.subheader("2. Resultados Calculados")

if not df_final.empty:
    st.dataframe(
        df_final[["Tiempo (s)", "Media Abs", "moles Fe2+", "q_n,p (einstein s⁻¹)"]].style.format({
            "Media Abs": "{:.4f}",
            "moles Fe2+": "{:.2e}",
            "q_n,p (einstein s⁻¹)": "{:.4e}",
        }),
        use_container_width=True,
    )
else:
    st.warning("No hay réplicas seleccionadas o datos disponibles.")

# 3. REGRESIÓN LINEAL (Absorbancia vs Tiempo)
st.markdown("---")
st.subheader("3. Parámetros de Regresión Lineal (Absorbancia vs Tiempo)")

if not df_final.empty and len(df_final) > 1:
    x_abs = df_final["Tiempo (s)"].values
    y_abs = df_final["Media Abs"].values
    slope_abs, intercept_abs, r_val_abs, _, _ = linregress(x_abs, y_abs)
    r_sq_abs = r_val_abs**2

    col_reg1, col_reg2, col_reg3 = st.columns(3)
    col_reg1.metric(label="Pendiente (ΔAbs/s)", value=f"{slope_abs:.4e}")
    col_reg2.metric(label="Intercepto (Abs₀)", value=f"{intercept_abs:.4f}")
    col_reg3.metric(label="Coeficiente de Correlación (R²)", value=f"{r_sq_abs:.4f}")
else:
    slope_abs, intercept_abs, r_sq_abs = 0.0, 0.0, 0.0
    st.info("Ingrese al menos 2 puntos de tiempo activos para calcular la regresión lineal.")

# Regresiones adicionales para otros gráficos
if not df_final.empty and len(df_final) > 1:
    slope_fe, intercept_fe, r_val_fe, _, _ = linregress(df_final["Tiempo (s)"], df_final["moles Fe2+"])
    r_sq_fe = r_val_fe**2
    
    df_q = df_final[df_final["Tiempo (s)"] > 0]
    if len(df_q) > 1:
        slope_q, intercept_q, r_val_q, _, _ = linregress(df_q["Tiempo (s)"], df_q["q_n,p (einstein s⁻¹)"])
        r_sq_q = r_val_q**2
    else:
        slope_q, intercept_q, r_sq_q = 0.0, 0.0, 0.0
else:
    slope_fe, intercept_fe, r_sq_fe = 0.0, 0.0, 0.0
    slope_q, intercept_q, r_sq_q = 0.0, 0.0, 0.0

# 4. GRÁFICOS INTERACTIVOS
st.markdown("---")
st.subheader("4. Gráficos Interactivos")

if not df_final.empty:
    # Gráfico 1: Absorbancia vs Tiempo
    fig_abs = go.Figure()
    fig_abs.add_trace(go.Scatter(
        x=df_final["Tiempo (s)"],
        y=df_final["Media Abs"],
        error_y=dict(type='data', array=df_final["Std Abs"], visible=True),
        mode='lines+markers',
        name='Promedio Absorbancia',
        line=dict(color='#2a9d8f', width=2)
    ))
    if len(df_final) > 1:
        fig_abs.add_trace(go.Scatter(
            x=df_final["Tiempo (s)"],
            y=intercept_abs + slope_abs * df_final["Tiempo (s)"],
            mode='lines',
            name=f"Ajuste lineal (R² = {r_sq_abs:.3f})",
            line=dict(color='orange', dash='dash')
        ))
    fig_abs.update_layout(
        title=f"Absorbancia (510 nm) vs Tiempo (R² = {r_sq_abs:.3f})",
        xaxis_title="Tiempo (s)",
        yaxis_title="Absorbancia",
        template="plotly_white"
    )
    st.plotly_chart(fig_abs, use_container_width=True)

    # Gráficos inferiores (2 columnas)
    gcol1, gcol2 = st.columns(2)

    with gcol1:
        fig1 = go.Figure()
        fig1.add_trace(go.Scatter(
            x=df_final["Tiempo (s)"],
            y=df_final["moles Fe2+"],
            error_y=dict(type='data', array=df_final["Std moles Fe2+"], visible=True),
            mode='lines+markers',
            name='moles Fe²⁺',
            line=dict(color='#1d3557', width=2)
        ))
        if len(df_final) > 1:
            fig1.add_trace(go.Scatter(
                x=df_final["Tiempo (s)"],
                y=intercept_fe + slope_fe * df_final["Tiempo (s)"],
                mode='lines',
                name=f"Ajuste lineal (R² = {r_sq_fe:.3f})",
                line=dict(color='red', dash='dash')
            ))
        fig1.update_layout(
            title=f"moles de Fe²⁺ (R² = {r_sq_fe:.3f})",
            xaxis_title="Tiempo (s)",
            yaxis_title="moles Fe²⁺",
            template="plotly_white"
        )
        st.plotly_chart(fig1, use_container_width=True)

    with gcol2:
        fig2 = go.Figure()
        fig2.add_trace(go.Scatter(
            x=df_final["Tiempo (s)"],
            y=df_final["q_n,p (einstein s⁻¹)"],
            error_y=dict(type='data', array=df_final["Std q_n,p"], visible=True),
            mode='lines+markers',
            name='qₙ,ₚ',
            line=dict(color='#e63946', width=2)
        ))
        if len(df_q) > 1:
            fig2.add_trace(go.Scatter(
                x=df_q["Tiempo (s)"],
                y=intercept_q + slope_q * df_q["Tiempo (s)"],
                mode='lines',
                name=f"Tendencia (R² = {r_sq_q:.3f})",
                line=dict(color='blue', dash='dash')
            ))
        fig2.update_layout(
            title=f"Evolución de qₙ,ₚ (R² = {r_sq_q:.3f})",
            xaxis_title="Tiempo (s)",
            yaxis_title="qₙ,ₚ (einstein s⁻¹)",
            template="plotly_white"
        )
        st.plotly_chart(fig2, use_container_width=True)

# EVALUADOR INTERACTIVO DE FLUJO FOTÓNICO (q_n,p)
st.markdown("---")
st.subheader("Evaluador Interactivo del Flujo Fotónico (qₙ,ₚ)")
st.write("Ingrese un valor de **Tiempo (en segundos)** para calcular el valor estimado del flujo fotónico **qₙ,ₚ (einstein s⁻¹)** utilizando la ecuación derivada de los datos activos:")

col_eval1, col_eval2 = st.columns([1, 2])
with col_eval1:
    x_usuario = st.number_input("Valor de Tiempo X (s)", value=10.0, format="%.2f", min_value=0.1)

q_calculado = slope_q * x_usuario + intercept_q if len(df_q) > 1 else 0.0

with col_eval2:
    if len(df_q) > 1:
        st.markdown(f"**Ecuación de Tendencia para qₙ,ₚ:**  \n$q_{{n,p}} = {slope_q:.4e} \\cdot X + ({intercept_q:.4e})$")
        st.markdown(f"**Coeficiente de Correlación (R²):** {r_sq_q:.4f}")
    st.success(f"Para un Tiempo de **{x_usuario} s**, el flujo fotónico estimado ($q_{{n,p}}$) es: **{q_calculado:.6e} einstein s⁻¹**")

# 5. SECCIÓN DE ECUACIONES
st.markdown("---")
st.subheader("5. Fundamento Teórico")

ecuaciones_path = os.path.join(os.path.dirname(__file__), "ecuaciones.png")
if os.path.exists(ecuaciones_path):
    st.image(ecuaciones_path, use_container_width=True)
elif os.path.exists("ecuaciones.png"):
    st.image("ecuaciones.png", use_container_width=True)
else:
    st.info(
        "Guarda la imagen con el nombre 'ecuaciones.png' en la carpeta de la"
        " aplicación para visualizarla aquí."
    )

# 6. EXPORTACIÓN INTEGRAL EN EXCEL CON FÓRMULAS Y LOS 3 GRÁFICOS
st.markdown("---")
st.subheader("6. Exportar Datos")

def generar_excel_con_formulas():
    output = io.BytesIO()
    wb = openpyxl.Workbook()

    ws_datos = wb.active
    ws_datos.title = "Datos Experimentales"

    headers_datos = list(df_editado.columns)
    ws_datos.append(headers_datos)
    for _, row in df_editado.iterrows():
        ws_datos.append(list(row))

    ws_res = wb.create_sheet(title="Resultados")

    ws_res["A1"] = "Constantes y Parámetros"
    ws_res["A2"], ws_res["B2"] = "V1 - Volumen reactor (L)", v1
    ws_res["A3"], ws_res["B3"] = "V2 - Volumen alícuota (L)", v2
    ws_res["A4"], ws_res["B4"] = "V3 - Volumen enrase (L)", v3
    ws_res["A5"], ws_res["B5"] = "V_total (L)", "=B2+B3+B4"
    ws_res["A6"], ws_res["B6"] = "Epsilon (M⁻¹cm⁻¹)", epsilon
    ws_res["A7"], ws_res["B7"] = "Paso óptico l (cm)", l
    ws_res["A8"], ws_res["B8"] = "Phi(lambda)", phi_lambda

    headers_res = [
        "Tiempo (s)",
        "Media Abs",
        "Moles Fe2+",
        "q_n,p (einstein s⁻¹)",
    ]
    for col_idx, h in enumerate(headers_res, 1):
        ws_res.cell(row=11, column=col_idx, value=h)

    num_filas = len(df_editado)
    for i in range(num_filas):
        row_idx = 12 + i
        src_row = 2 + i

        ws_res.cell(row=row_idx, column=1, value=f"='Datos Experimentales'!A{src_row}")
        ws_res.cell(
            row=row_idx,
            column=2,
            value=f"=AVERAGE('Datos Experimentales'!B{src_row}:D{src_row})",
        )
        ws_res.cell(
            row=row_idx, column=3, value=f"=(B{row_idx} * $B$5) / ($B$6 * $B$7)"
        )
        if i == 0:
            ws_res.cell(row=row_idx, column=4, value=0)
        else:
            ws_res.cell(
                row=row_idx,
                column=4,
                value=(
                    f"=(B{row_idx} * $B$2 * $B$4) / ($B$8 * $B$6 * $B$3 * $B$7 *"
                    f" A{row_idx})"
                ),
            )

    xvalues = Reference(ws_res, min_col=1, min_row=12, max_row=11 + num_filas)

    # Gráfico 1: Absorbancia vs Tiempo
    chart_abs = ScatterChart()
    chart_abs.title = "Absorbancia vs Tiempo"
    chart_abs.style = 13
    chart_abs.x_axis.title = "Tiempo (s)"
    chart_abs.y_axis.title = "Absorbancia"
    values_abs = Reference(ws_res, min_col=2, min_row=11, max_row=11 + num_filas)
    series_abs = Series(values_abs, xvalues=xvalues, title_from_data=True)
    chart_abs.series.append(series_abs)
    ws_res.add_chart(chart_abs, "F2")

    # Gráfico 2: Moles Fe2+ vs Tiempo
    chart1 = ScatterChart()
    chart1.title = "moles de Fe²⁺"
    chart1.style = 13
    chart1.x_axis.title = "Tiempo (s)"
    chart1.y_axis.title = "moles Fe²⁺"
    values1 = Reference(ws_res, min_col=3, min_row=11, max_row=11 + num_filas)
    series1 = Series(values1, xvalues=xvalues, title_from_data=True)
    chart1.series.append(series1)
    ws_res.add_chart(chart1, "F17")

    # Gráfico 3: q_n,p vs Tiempo
    chart2 = ScatterChart()
    chart2.title = "Evolución de qₙ,ₚ"
    chart2.style = 13
    chart2.x_axis.title = "Tiempo (s)"
    chart2.y_axis.title = "qₙ,ₚ (einstein s⁻¹)"
    values2 = Reference(ws_res, min_col=4, min_row=11, max_row=11 + num_filas)
    series2 = Series(values2, xvalues=xvalues, title_from_data=True)
    chart2.series.append(series2)
    ws_res.add_chart(chart2, "F32")

    wb.save(output)
    return output.getvalue()

excel_data = generar_excel_con_formulas()

st.download_button(
    label="📥 Descargar Resultados en Excel",
    data=excel_data,
    file_name="Resultados_flujo_fotonico.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

# Versión
st.markdown("---")
st.markdown(
    "<div style='text-align: center; color: gray; font-size: 14px;'>"
    "Desarrollado por Cristian Poza Retamales. v1.0_13.08.2026"
    "</div>",
    unsafe_allow_html=True,
)